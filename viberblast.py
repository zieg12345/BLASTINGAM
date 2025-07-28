import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import random
import openpyxl
from openpyxl import Workbook
from io import BytesIO

# Initialize session state
if 'button1_clicked' not in st.session_state:
    st.session_state.button1_clicked = False
if 'button2_clicked' not in st.session_state:
    st.session_state.button2_clicked = False
if 'button3_clicked' not in st.session_state:
    st.session_state.button3_clicked = False
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'collector_file' not in st.session_state:
    st.session_state.collector_file = None
if 'menu_open' not in st.session_state:
    st.session_state.menu_open = False

# Set page configuration
st.set_page_config(page_title="WORKLOADS-AUTOMATED", page_icon="📊", layout="wide")

# Custom CSS
st.markdown(
    """
    <style>
    .main-content {
        padding: 20px;
        background-color: #f5f5f5;
        border-radius: 10px;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        color: #2b2b2b;
        margin-bottom: 20px;
    }
    .css-1lcbmhc { 
        background-color: #e0e0e0;
        padding: 20px;
        border-radius: 10px;
        transition: all 0.3s ease;
        width: 250px;
    }
    .stButton > button {
        width: 100%;
        margin-bottom: 10px;
        padding: 12px;
        background-color: #b0b0b0;
        color: #2b2b2b;
        border: none;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
        transition: all 0.2s ease;
        font-size: 16px;
        font-weight: 500;
        text-align: center;
    }
    .stButton > button:hover {
        background-color: #909090;
        transform: scale(1.03);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.3);
    }
    .stDownloadButton > button {
        background-color: #b0b0b0;
        color: #2b2b2b;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
        transition: all 0.2s ease;
        font-size: 16px;
        font-weight: 500;
        padding: 12px;
        width: 100%;
    }
    .stDownloadButton > button:hover {
        background-color: #909090;
        transform: scale(1.03);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.3);
    }
    h1 {
        text-align: center;
        color: #2b2b2b;
        font-size: 24px;
        font-weight: 600;
    }
    .burger-button {
        font-size: 24px;
        cursor: pointer;
        color: #2b2b2b;
        margin-bottom: 15px;
        background: none;
        border: none;
        padding: 5px;
        transition: all 0.2s ease;
    }
    .burger-button:hover {
        color: #606060;
    }
    .sidebar-content {
        display: none;
    }
    .sidebar-content.active {
        display: block;
    }
    .stDataFrame {
        border: 1px solid #d0d0d0;
        border-radius: 5px;
        background-color: #ffffff;
    }
    .dashboard-iframe {
        border: none;
        border-radius: 10px;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        width: 100%;
        height: 600px;
    }
    .footer {
        text-align: center;
        color: #666666;
        margin-top: 20px;
        font-size: 12px;
    }
    @media (max-width: 768px) {
        .css-1lcbmhc {
            width: 100%;
            padding: 10px;
        }
        .stButton > button {
            font-size: 14px;
            padding: 10px;
        }
        .stDownloadButton > button {
            font-size: 14px;
            padding: 10px;
        }
        h1 {
            font-size: 20px;
        }
        .dashboard-iframe {
            height: 400px;
        }
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Motivational quotes
motivational_quotes = [
    "Trust in your inner strength—you’ve already crossed half the journey. – Zieg",
    "Exceptional work blooms from a heart that loves its craft. – Zieg",
    "True success lies in the bravery to press forward despite challenges. – Zieg",
    "The boundaries you see are merely shadows of your own imagination. – Zieg",
    "Set grand goals, toil relentlessly, remain steadfast, and choose wise companions. – Zieg",
    "Tomorrow is crafted by those who envision their dreams with wonder. – Zieg",
    "Don’t follow the ticking clock—mirror its persistence and keep advancing. – Zieg",
    "No age can stop you from pursuing new dreams or crafting fresh ambitions. – Zieg",
    "The greatest prize of your achievements is the person you grow into. – Zieg",
    "Launch from your current place, with your present tools, and give your all. – Zieg"
]

# Select a random quote
random_quote = random.choice(motivational_quotes)

# Header section
col1, col2, col3 = st.columns([1, 3, 1])
with col2:
    st.title(random_quote)

# Sidebar with burger menu
with st.sidebar:
    if st.session_state.menu_open:
        if st.button("✕ Close", key="close_menu", help="Close the menu"):
            st.session_state.menu_open = False
    else:
        if st.button("☰", key="burger_menu", help="Open the menu"):
            st.session_state.menu_open = True

    if st.session_state.menu_open:
        st.markdown('<div class="sidebar-content active">', unsafe_allow_html=True)
        if st.button("VIBER BLAST", help="Access Viber Blast CSV Uploader"):
            st.session_state.button1_clicked = True
            st.session_state.button2_clicked = False
            st.session_state.button3_clicked = False
            st.session_state.uploaded_file = None
            st.session_state.collector_file = None
        if st.button("EMAIL BLAST", help="Access Email Blast File Uploader"):
            st.session_state.button1_clicked = False
            st.session_state.button2_clicked = True
            st.session_state.button3_clicked = False
            st.session_state.uploaded_file = None
            st.session_state.collector_file = None
        if st.button("LIVE INBOUND MONITORING", help="Access MC4 Blasting Monitoring Dashboard"):
            st.session_state.button1_clicked = False
            st.session_state.button2_clicked = False
            st.session_state.button3_clicked = True
            st.session_state.uploaded_file = None
            st.session_state.collector_file = None
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-content">', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

# Main content
with st.container():
    st.markdown('<div class="main-content">', unsafe_allow_html=True)
    
    if not (st.session_state.button1_clicked or st.session_state.button2_clicked or st.session_state.button3_clicked):
        st.subheader("Welcome")
        st.write("Click the ☰ menu in the sidebar and select a feature to begin.")
    elif st.session_state.button1_clicked:
        st.subheader("Viber Blast CSV Uploader")
        
        # Dropdown for selecting bucket
        bucket_option = st.selectbox(
            "Select Campaign",
            ["Bucket 2", "Bucket 4"],
            help="Choose the bucket for Viber blast processing"
        )

        # File uploader
        uploaded_file = st.file_uploader(
            "📤 Choose a CSV file",
            type=["csv"],
            key=f"viber_blast_uploader_{bucket_option.lower().replace(' ', '_')}",
            help="Upload a CSV with columns: Client, Account No., Debtor Name, Contact No."
        )
        if uploaded_file is not None:
            st.session_state.uploaded_file = uploaded_file
            st.success("File uploaded successfully!")

        # Reset button
        if st.session_state.uploaded_file is not None:
            if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                st.session_state.uploaded_file = None
                st.session_state.button1_clicked = False
                st.rerun()

        # Sample data based on bucket
        if bucket_option == "Bucket 2":
            sample_data = {
                "Campaign": ["SAMPLE", "SAMPLE", "SAMPLE", "SAMPLE"],
                "CH Code": ["12345", "123456", "1234567", "12345678"],
                "First Name": ["", "", "", ""],
                "Full Name": ["Richard Arenas", "Jinnggoy Dela Cruz", "Roman Dalisay", "Edwin Paras"],
                "Last Name": ["", "", "", ""],
                "Mobile Number": ["09274186327", "09760368821", "09088925110", "09175791122"],
                "OB": ["", "", "", ""]
            }
        else:  # Bucket 4
            sample_data = {
                "Campaign": ["SAMPLE"],
                "CH Code": ["123456789"],
                "First Name": [""],
                "Full Name": ["Janica d Benbinuto"],
                "Last Name": [""],
                "Mobile Number": ["09655669672"],
                "OB": [""]
            }
        sample_df = pd.DataFrame(sample_data)

        # Dynamic filename
        current_date = datetime.now().strftime(f"VIBER BLAST {bucket_option.upper()} %b %d %Y %I:%M %p PST").upper()

        if st.session_state.uploaded_file is not None:
            try:
                df = pd.read_csv(st.session_state.uploaded_file, encoding='utf-8-sig', skipinitialspace=True)
                df.columns = df.columns.str.strip()
                required_columns = ["Client", "Account No.", "Debtor Name", "Contact No."]
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    st.error(f"The following required columns are missing: {', '.join(missing_columns)}")
                else:
                    df["Contact No."] = df["Contact No."].astype(str).str.strip().replace("nan", "")
                    df["Account No."] = df["Account No."].astype(str).str.strip().replace("nan", "")
                    invalid_contact_no = df[df["Contact No."].str.len() != 11]
                    if not invalid_contact_no.empty:
                        st.warning(f"Found {len(invalid_contact_no)} rows where Contact No. is not 11 digits. These rows are still included but may need review.")
                    initial_row_count_bel = len(df)
                    df = df[~df["Account No."].str.contains("BEL", case=False, na=False)]
                    if initial_row_count_bel != len(df):
                        st.info(f"Removed {initial_row_count_bel - len(df)} rows where Account No. contains 'BEL'.")
                    initial_row_count = len(df)
                    df = df.drop_duplicates(subset=["Account No."], keep="first")
                    if initial_row_count != len(df):
                        st.info(f"Removed {initial_row_count - len(df)} duplicate rows based on 'Account No.'.")
                    if len(df) == 0:
                        st.warning("No rows remain after filtering. Showing sample data only.")
                    summary_df = pd.DataFrame({
                        "Campaign": df["Client"],
                        "CH Code": df["Account No."],
                        "First Name": [""] * len(df),
                        "Full Name": df["Debtor Name"],
                        "Last Name": [""] * len(df),
                        "Mobile Number": df["Contact No."],
                        "OB": [""] * len(df)
                    })
                    summary_df["Mobile Number"] = summary_df["Mobile Number"].astype(str)
                    summary_df["CH Code"] = summary_df["CH Code"].astype(str)
                    summary_df = pd.concat([summary_df, sample_df], ignore_index=True)
                    st.subheader("Summary Table")
                    st.dataframe(summary_df, use_container_width=True)
                    output = BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Viber Blast"
                    headers = list(summary_df.columns)
                    for col_num, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_num).value = header
                    for row_num, row in enumerate(summary_df.values, 2):
                        for col_num, value in enumerate(row, 1):
                            ws.cell(row=row_num, column=col_num).value = value
                            if headers[col_num-1] in ["Mobile Number", "CH Code"]:
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                    wb.save(output)
                    output.seek(0)
                    st.download_button(
                        label="📥 Download Summary Table as Excel",
                        data=output,
                        file_name=f"{current_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_summary"
                    )
            except Exception as e:
                st.error(f"An error occurred while processing the file: {str(e)}")
        else:
            st.subheader("Sample Summary Table")
            st.dataframe(sample_df, use_container_width=True)
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Viber Blast"
            headers = list(sample_df.columns)
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            for row_num, row in enumerate(sample_df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num).value = value
                    if headers[col_num-1] in ["Mobile Number", "CH Code"]:
                        ws.cell(row=row_num, column=col_num).number_format = '@'
            wb.save(output)
            output.seek(0)
            st.download_button(
                label="📥 Download",
                data=output,
                file_name=f"{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_sample"
            )
            st.info("Please upload a CSV file to generate the summary table with your data.")
    elif st.session_state.button2_clicked:
        st.subheader("Email Blast File Uploader")
        bucket_option = st.selectbox(
            "Select Campaign",
            ["Bucket 2 with sequence template", "Bucket 4 Generic Template", "LEVEL 1 NEGATIVE ACCOUNTS", 
             "LEVEL 6 NEGATIVE ACCOUNTS", "SBF SALAD NEGATIVE ACCOUNT", "SBF PL NEGATIVE ACCOUNTS"],
            help="Choose the bucket for email blast processing"
        )

        if bucket_option == "Bucket 2 with sequence template":
            uploaded_file = st.file_uploader(
                "📤 Choose a CSV or Excel file",
                type=["csv", "xlsx"],
                key="email_blast_uploader",
                help="Upload a CSV or Excel (.xlsx) file with columns: Contract Number, Email, {{chname}}, Statement Balance (OB), Statement Overdue Amount (MYP), Statement Minimum Payment (MAD), Assignment Date, TEMPLATE 1 D1, TEMPLATE 1 D2, etc."
            )
            if uploaded_file is not None:
                st.session_state.uploaded_file = uploaded_file
                st.success("Main file uploaded successfully!")
            collector_file = st.file_uploader(
                "📤 Choose a CSV or Excel file for Collector and Assign Date data",
                type=["csv", "xlsx"],
                key="collector_uploader",
                help="Upload a CSV or Excel (.xlsx) file with columns: Financing/Card No., Collector, Assign Date"
            )
            if collector_file is not None:
                st.session_state.collector_file = collector_file
                st.success("Collector file uploaded successfully!")
            if st.session_state.uploaded_file is not None or st.session_state.collector_file is not None:
                if st.button("🔄 Reset", help="Clear the uploaded files and reset"):
                    st.session_state.uploaded_file = None
                    st.session_state.collector_file = None
                    st.session_state.button2_clicked = False
                    st.rerun()
            if st.session_state.uploaded_file is not None:
                try:
                    if st.session_state.uploaded_file.name.endswith('.csv'):
                        df = pd.read_csv(st.session_state.uploaded_file, encoding='utf-8-sig', skipinitialspace=True)
                    elif st.session_state.uploaded_file.name.endswith('.xlsx'):
                        df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl')
                    if df.empty:
                        st.error("The uploaded file is empty. Please upload a valid file.")
                        st.stop()
                    df.columns = df.columns.str.strip()
                    # Validate Email column for '@' symbol
                    initial_row_count_email = len(df)
                    df['Email'] = df['Email'].astype(str)
                    df = df[df['Email'].str.contains('@', na=False)]
                    if len(df) < initial_row_count_email:
                        st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
                    required_columns = [
                        "Contract Number", "Email", "{{chname}}", "Statement Balance (OB)",
                        "Statement Overdue Amount (MYP)", "Statement Minimum Payment (MAD)",
                        "Assignment Date", "TEMPLATE 1 D1", "TEMPLATE 1 D2", "TEMPLATE 2 D1",
                        "TEMPLATE 2 D2", "TEMPLATE 3 D1", "TEMPLATE 3 D2", "TEMPLATE 4 D1",
                        "TEMPLATE 4 D2", "TEMPLATE 5 D1", "TEMPLATE 5 D2", "TEMPLATE 6 D1",
                        "TEMPLATE 6 D2"
                    ]
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        st.error(f"The following required columns are missing in the main file: {', '.join(missing_columns)}")
                        st.stop()
                    initial_row_count = len(df)
                    df = df.drop_duplicates(subset="Contract Number", keep="first")
                    if initial_row_count != len(df):
                        st.info(f"Removed {initial_row_count - len(df)} duplicate rows based on 'Contract Number'.")
                    with st.expander("🔍 Show Detected Column Names (Main File)"):
                        st.write("Detected Column Names:", list(df.columns))
                    summary_df = pd.DataFrame()
                    summary_df["Contract Number"] = df["Contract Number"].astype(str).str.replace(r'\.0$', '', regex=True)
                    summary_df["Email"] = df["Email"]
                    summary_df["{{chname}}"] = df["{{chname}}"]
                    summary_df["{{agentcode}}"] = ""
                    summary_df["{{ID}}"] = ""
                    summary_df["{{OB}}"] = pd.to_numeric(df["Statement Balance (OB)"], errors='coerce').apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
                    summary_df["{{MYP}}"] = pd.to_numeric(df["Statement Overdue Amount (MYP)"], errors='coerce').apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
                    summary_df["{{MAD}}"] = pd.to_numeric(df["Statement Minimum Payment (MAD)"], errors='coerce').apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
                    summary_df["{{OB+CF}}"] = pd.to_numeric(df["Statement Balance (OB)"], errors='coerce') * 1.11
                    summary_df["{{OB+CF}}"] = summary_df["{{OB+CF}}"].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
                    summary_df["{{MAD+CF}}"] = pd.to_numeric(df["Statement Minimum Payment (MAD)"], errors='coerce') * 1.11
                    summary_df["{{MAD+CF}}"] = summary_df["{{MAD+CF}}"].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
                    summary_df["{{MYP+CF}}"] = pd.to_numeric(df["Statement Overdue Amount (MYP)"], errors='coerce') * 1.11
                    summary_df["{{MYP+CF}}"] = summary_df["{{MYP+CF}}"].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")
                    summary_df["TEMPLATE 1 D1"] = df["TEMPLATE 1 D1"]
                    summary_df["TEMPLATE 1 D2"] = df["TEMPLATE 1 D2"]
                    summary_df["TEMPLATE 2 D1"] = df["TEMPLATE 2 D1"]
                    summary_df["TEMPLATE 2 D2"] = df["TEMPLATE 2 D2"]
                    summary_df["TEMPLATE 3 D1"] = df["TEMPLATE 3 D1"]
                    summary_df["TEMPLATE 3 D2"] = df["TEMPLATE 3 D2"]
                    summary_df["TEMPLATE 4 D1"] = df["TEMPLATE 4 D1"]
                    summary_df["TEMPLATE 4 D2"] = df["TEMPLATE 4 D2"]
                    summary_df["TEMPLATE 5 D1"] = df["TEMPLATE 5 D1"]
                    summary_df["TEMPLATE 5 D2"] = df["TEMPLATE 5 D2"]
                    summary_df["TEMPLATE 6 D1"] = df["TEMPLATE 6 D1"]
                    summary_df["TEMPLATE 6 D2"] = df["TEMPLATE 6 D2"]
                    if st.session_state.collector_file is not None:
                        try:
                            if st.session_state.collector_file.name.endswith('.csv'):
                                collector_df = pd.read_csv(st.session_state.collector_file, encoding='utf-8-sig', skipinitialspace=True)
                            elif st.session_state.collector_file.name.endswith('.xlsx'):
                                collector_df = pd.read_excel(st.session_state.collector_file, engine='openpyxl')
                            with st.expander("🔍 Show Detected Column Names (Collector File)"):
                                st.write("Detected Column Names:", list(collector_df.columns))
                            collector_required_columns = ["Financing/Card No.", "Collector", "Assign Date"]
                            collector_missing_columns = [col for col in collector_required_columns if col not in collector_df.columns]
                            if collector_missing_columns:
                                st.error(f"Missing required columns in collector file: {', '.join(collector_missing_columns)}")
                                st.stop()
                            collector_df["Financing/Card No."] = collector_df["Financing/Card No."].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                            summary_df["Contract Number"] = summary_df["Contract Number"].astype(str).str.strip()
                            try:
                                collector_df["Assign Date"] = pd.to_datetime(collector_df["Assign Date"], errors='coerce')
                            except:
                                st.error("Unable to parse 'Assign Date' column in collector file. Ensure dates are in a valid format (e.g., YYYY-MM-DD).")
                                st.stop()
                            summary_df = summary_df.merge(
                                collector_df[["Financing/Card No.", "Collector", "Assign Date"]],
                                how="left",
                                left_on="Contract Number",
                                right_on="Financing/Card No."
                            )
                            summary_df["{{agentcode}}"] = summary_df["Collector"].fillna("")
                            summary_df["Assignment Date"] = summary_df["Assign Date"].dt.strftime('%m/%d/%Y').fillna("")
                            summary_df = summary_df.drop(columns=["Financing/Card No.", "Collector", "Assign Date"], errors='ignore')
                            summary_df['{{agentcode}}'] = summary_df['{{agentcode}}'].apply(lambda x: 'PJHA' if x == 'SPMADRID' else x)
                            summary_df['{{ID}}'] = summary_df['{{agentcode}}'].apply(lambda x: 'BDCO' if x == 'PJHA' else 'BCCO' if x else '')
                            initial_row_count = len(summary_df)
                            summary_df = summary_df[summary_df["{{agentcode}}"].notna() & (summary_df["{{agentcode}}"] != "")]
                            if len(summary_df) < initial_row_count:
                                st.info(f"Removed {initial_row_count - len(summary_df)} rows where {{agentcode}} was blank or null.")
                            if summary_df["{{agentcode}}"].isna().any() or (summary_df["{{agentcode}}"] == "").any() or \
                               summary_df["Assignment Date"].isna().any() or (summary_df["Assignment Date"] == "").any():
                                st.warning("Some Contract Numbers did not match with Financing/Card No. in the collector file.")
                        except Exception as e:
                            st.error(f"An error occurred while processing the collector file: {str(e)}")
                    columns_to_check = [
                        "Email", "{{chname}}", "{{OB}}", "{{MYP}}", "{{MAD}}",
                        "{{OB+CF}}", "{{MAD+CF}}", "{{MYP+CF}}"
                    ]
                    summary_df = summary_df.dropna(subset=columns_to_check)
                    summary_df = summary_df[~(summary_df[columns_to_check] == "").any(axis=1)]
                    text_columns = [
                        "Contract Number", "Email", "{{chname}}", "{{agentcode}}", "{{ID}}",
                        "Assignment Date", "TEMPLATE 1 D1", "TEMPLATE 1 D2", "TEMPLATE 2 D1",
                        "TEMPLATE 2 D2", "TEMPLATE 3 D1", "TEMPLATE 3 D2", "TEMPLATE 4 D1",
                        "TEMPLATE 4 D2", "TEMPLATE 5 D1", "TEMPLATE 5 D2",
                        "TEMPLATE 6 D1", "TEMPLATE 6 D2"
                    ]
                    for col in text_columns:
                        if col in summary_df.columns:
                            summary_df[col] = summary_df[col].astype(str)
                    st.subheader("Summary Table")
                    if not summary_df.empty:
                        st.dataframe(summary_df, use_container_width=True)
                        st.markdown("<br>", unsafe_allow_html=True)
                        output = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Summary"
                        headers = list(summary_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(summary_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                                if headers[col_num-1] in text_columns:
                                    ws.cell(row=row_num, column=col_num).number_format = '@'
                        wb.save(output)
                        output.seek(0)
                        today = datetime.now().strftime("%B %d %Y")
                        file_name = f"B2 Email blasting {today}.xlsx"
                        st.download_button(
                            label="📥 Download Summary Table as Excel Workbook",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_email_summary",
                            use_container_width=True
                        )
                    else:
                        st.warning("No rows remain after removing those with blank or None in Email, {{chname}}, {{OB}}, {{MYP}}, {{MAD}}, {{OB+CF}}, {{MAD+CF}}, or {{MYP+CF}} fields.")
                except Exception as e:
                    st.error(f"An error occurred while processing the main file: {str(e)}")
            else:
                st.info("Please upload a CSV or Excel file to generate the summary table.")
        elif bucket_option == "Bucket 4 Generic Template":
            st.subheader("Bucket 4 Generic Template Email Blast File Uploader")
            uploaded_file = st.file_uploader(
                "📤 Choose an Excel file",
                type=["xlsx"],
                key="bucket4_uploader",
                help="Upload an Excel (.xlsx) file with columns: Email, Name, Collector, Product Type, Financing/Card No., Account No., Assign Date"
            )
            if uploaded_file is not None:
                st.session_state.uploaded_file = uploaded_file
                st.success("File uploaded successfully!")
            if st.session_state.uploaded_file is not None:
                if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                    st.session_state.uploaded_file = None
                    st.rerun()
            # Sample data for Bucket 4 Generic Template
            sample_data = {
                "Email": ["JDBenbinuto@securitybank.com.ph"],
                "{{chname}}": ["Janica d Benbinuto"],
                "{{agentcode}}": ["PJHA"],
                "{{product}}": ["CARD"],
                "Financing/Card No.": ["123456789"],
                "Account No.": ["987654321"],
                "Assign Date": [datetime.now().strftime('%Y-%m-%d')],
                "{{ID}}": ["4DCO"]
            }
            sample_df = pd.DataFrame(sample_data)
            if st.session_state.uploaded_file is not None:
                try:
                    df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl')
                    # Validate Email column for '@' symbol
                    initial_row_count_email = len(df)
                    df['Email'] = df['Email'].astype(str)
                    df = df[df['Email'].str.contains('@', na=False)]
                    if len(df) < initial_row_count_email:
                        st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
                    required_columns = ['Email', 'Name', 'Collector', 'Product Type', 'Financing/Card No.', 'Account No.', 'Assign Date']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        st.error(f"Missing required columns in the uploaded file: {', '.join(missing_columns)}")
                    else:
                        df = df[required_columns]
                        column_mapping = {
                            'Name': '{{chname}}',
                            'Collector': '{{agentcode}}',
                            'Product Type': '{{product}}'
                        }
                        df = df.rename(columns=column_mapping)
                        df['{{product}}'] = df['{{product}}'].replace({'MC': 'CARD', 'BEL': 'BUSINESS EXPRESS LOAN'})
                        try:
                            df['Assign Date'] = pd.to_datetime(df['Assign Date'])
                        except:
                            st.error("Unable to parse 'Assign Date' column in the file. Please ensure dates are in a valid format (e.g., YYYY-MM-DD).")
                            st.stop()
                        df['{{agentcode}}'] = df['{{agentcode}}'].apply(lambda x: 'PJHA' if x == 'SPMADRID' else x)
                        df['{{ID}}'] = df['{{agentcode}}'].apply(lambda x: '4DCO' if x == 'PJHA' else '4CCO')
                        text_columns = ['Email', '{{chname}}', '{{product}}', '{{agentcode}}', '{{ID}}', 'Financing/Card No.', 'Account No.']
                        for col in text_columns:
                            if col in df.columns:
                                df[col] = df[col].astype(str)
                        # Combine with sample data
                        summary_df = pd.concat([df, sample_df], ignore_index=True)
                        st.write("### Processed Data")
                        st.dataframe(summary_df, use_container_width=True)
                        st.write("### Summary")
                        total_records = len(summary_df)
                        unique_emails = summary_df['Email'].nunique()
                        unique_names = summary_df['{{chname}}'].nunique()
                        unique_agents = summary_df['{{agentcode}}'].nunique()
                        unique_products = summary_df['{{product}}'].nunique()
                        unique_accounts = summary_df['Account No.'].nunique()
                        unique_ids = summary_df['{{ID}}'].nunique()
                        try:
                            date_range = f"From {summary_df['Assign Date'].min().strftime('%Y-%m-%d')} to {summary_df['Assign Date'].max().strftime('%Y-%m-%d')}"
                        except:
                            date_range = "Invalid date format"
                        st.write(f"- **Total Records**: {total_records}")
                        st.write(f"- **Unique Emails**: {unique_emails}")
                        st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                        st.write(f"- **Unique Agents ({{agentcode}})**: {unique_agents}")
                        st.write(f"- **Unique Products**: {unique_products}")
                        st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                        st.write(f"- **Unique IDs ({{ID}})**: {unique_ids}")
                        st.write(f"- **Assign Date Range**: {date_range}")
                        output = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Summary"
                        headers = list(summary_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(summary_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                                if headers[col_num-1] in text_columns:
                                    ws.cell(row=row_num, column=col_num).number_format = '@'
                        wb.save(output)
                        output.seek(0)
                        today = datetime.now().strftime("%B %d %Y")
                        file_name = f"B4 Email blasting {today}.xlsx"
                        st.download_button(
                            label="📥 Download Processed Excel",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_bucket4_summary",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"An error occurred while processing the file: {str(e)}")
            else:
                st.subheader("Sample Summary Table")
                st.dataframe(sample_df, use_container_width=True)
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Summary"
                headers = list(sample_df.columns)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num).value = header
                for row_num, row in enumerate(sample_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num).value = value
                        if headers[col_num-1] in ['Email', '{{chname}}', '{{agentcode}}', '{{product}}', '{{ID}}', 'Financing/Card No.', 'Account No.']:
                            ws.cell(row=row_num, column=col_num).number_format = '@'
                wb.save(output)
                output.seek(0)
                today = datetime.now().strftime("%B %d %Y")
                file_name = f"B4 Email blasting {today}.xlsx"
                st.download_button(
                    label="📥 Download Sample Excel",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_bucket4_sample",
                    use_container_width=True
                )
                st.info("Please upload an Excel file to generate the summary table with your data.")
        elif bucket_option == "LEVEL 1 NEGATIVE ACCOUNTS":
            st.subheader("LEVEL 1 NEGATIVE ACCOUNTS Email Blast File Uploader")
            uploaded_file = st.file_uploader(
                "📤 Choose an Excel file",
                type=["xlsx"],
                key="level1_uploader",
                help="Upload an Excel (.xlsx) with columns: Email, Name, Product Type, Client Name, Account No., Financing/Card No."
            )
            if uploaded_file is not None:
                st.session_state.uploaded_file = uploaded_file
                st.success("File uploaded successfully!")
            if st.session_state.uploaded_file is not None:
                if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                    st.session_state.uploaded_file = None
                    st.rerun()
            if st.session_state.uploaded_file is not None:
                try:
                    df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl')
                    # Validate Email column for '@' symbol
                    initial_row_count_email = len(df)
                    df['Email'] = df['Email'].astype(str)
                    df = df[df['Email'].str.contains('@', na=False)]
                    if len(df) < initial_row_count_email:
                        st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
                    required_columns = ['Email', 'Name', 'Product Type', 'Client Name', 'Account No.', 'Financing/Card No.']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        st.error(f"Missing required columns in the uploaded file: {', '.join(missing_columns)}")
                    else:
                        df = df[required_columns]
                        summary_df = pd.DataFrame({
                            'Email': df['Email'],
                            '{{chname}}': df['Name'],
                            '{{product}}': df['Product Type'],
                            '{{agentcode}}': 'PJND',
                            'Client Name': df['Client Name'],
                            'Account No.': df['Account No.'],
                            'Financing/Card No.': df['Financing/Card No.']
                        })
                        text_columns = ['Email', '{{chname}}', '{{product}}', '{{agentcode}}', 'Client Name', 'Account No.', 'Financing/Card No.']
                        for col in text_columns:
                            summary_df[col] = summary_df[col].astype(str)
                        st.write("### Processed Data")
                        st.dataframe(summary_df, use_container_width=True)
                        st.write("### Summary")
                        st.write("Note: Values in Account No. and Financing/Card No. are preserved exactly as uploaded.")
                        total_records = len(summary_df)
                        unique_emails = summary_df['Email'].nunique()
                        unique_names = summary_df['{{chname}}'].nunique()
                        unique_products = summary_df['{{product}}'].nunique()
                        unique_agents = summary_df['{{agentcode}}'].nunique()
                        unique_clients = summary_df['Client Name'].nunique()
                        unique_accounts = summary_df['Account No.'].nunique()
                        unique_financing = summary_df['Financing/Card No.'].nunique()
                        st.write(f"- **Total Records**: {total_records}")
                        st.write(f"- **Unique Emails**: {unique_emails}")
                        st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                        st.write(f"- **Unique Products**: {unique_products}")
                        st.write(f"- **Unique Agents ({{agentcode}})**: {unique_agents}")
                        st.write(f"- **Unique Client Names**: {unique_clients}")
                        st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                        st.write(f"- **Unique Financing/Card Numbers**: {unique_financing}")
                        output = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Summary"
                        headers = list(summary_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(summary_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                        wb.save(output)
                        output.seek(0)
                        today = datetime.now().strftime("%B %d %Y")
                        file_name = f"Level 1 Negative Accounts Email blasting {today}.xlsx"
                        st.download_button(
                            label="📥 Download Processed Excel",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_level1_summary",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"An error occurred while processing the file: {str(e)}")
            else:
                st.info("Please upload an Excel file to generate the summary table.")
        elif bucket_option == "LEVEL 6 NEGATIVE ACCOUNTS":
            st.subheader("LEVEL 6 NEGATIVE ACCOUNTS Email Blast File Uploader")
            uploaded_file = st.file_uploader(
                "📤 Choose an Excel file",
                type=["xlsx"],
                key="level6_uploader",
                help="Upload an Excel (.xlsx) with columns: Email, Name, Product Type, Client Name, Account No., Financing/Card No."
            )
            if uploaded_file is not None:
                st.session_state.uploaded_file = uploaded_file
                st.success("File uploaded successfully!")
            if st.session_state.uploaded_file is not None:
                if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                    st.session_state.uploaded_file = None
                    st.rerun()
            if st.session_state.uploaded_file is not None:
                try:
                    df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl')
                    # Validate Email column for '@' symbol
                    initial_row_count_email = len(df)
                    df['Email'] = df['Email'].astype(str)
                    df = df[df['Email'].str.contains('@', na=False)]
                    if len(df) < initial_row_count_email:
                        st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
                    required_columns = ['Email', 'Name', 'Product Type', 'Client Name', 'Account No.', 'Financing/Card No.']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        st.error(f"Missing required columns: {', '.join(missing_columns)}")
                    else:
                        df = df[required_columns]
                        summary_df = pd.DataFrame({
                            'Email': df['Email'],
                            '{{chname}}': df['Name'],
                            '{{product}}': df['Product Type'],
                            '{{agentcode}}': 'PJND6',
                            'Client Name': df['Client Name'],
                            'Account No.': df['Account No.'],
                            'Financing/Card No.': df['Financing/Card No.']
                        })
                        text_columns = ['Email', '{{chname}}', '{{product}}', '{{agentcode}}', 'Client Name', 'Account No.', 'Financing/Card No.']
                        for col in text_columns:
                            summary_df[col] = summary_df[col].astype(str)
                        st.write("### Processed Data")
                        st.dataframe(summary_df, use_container_width=True)
                        st.write("### Summary")
                        st.write("Note: Values in Account No. and Financing/Card No. are preserved exactly as uploaded.")
                        total_records = len(summary_df)
                        unique_emails = summary_df['Email'].nunique()
                        unique_names = summary_df['{{chname}}'].nunique()
                        unique_products = summary_df['{{product}}'].nunique()
                        unique_agents = summary_df['{{agentcode}}'].nunique()
                        unique_clients = summary_df['Client Name'].nunique()
                        unique_accounts = summary_df['Account No.'].nunique()
                        unique_financing = summary_df['Financing/Card No.'].nunique()
                        st.write(f"- **Total Records**: {total_records}")
                        st.write(f"- **Unique Emails**: {unique_emails}")
                        st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                        st.write(f"- **Unique Products**: {unique_products}")
                        st.write(f"- **Unique Agents ({{agentcode}})**: {unique_agents}")
                        st.write(f"- **Unique Client Names**: {unique_clients}")
                        st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                        st.write(f"- **Unique Financing/Card Numbers**: {unique_financing}")
                        output = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Summary"
                        headers = list(summary_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(summary_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                        wb.save(output)
                        output.seek(0)
                        today = datetime.now().strftime("%B %d %Y")
                        file_name = f"Level 6 Negative Accounts Email blasting {today}.xlsx"
                        st.download_button(
                            label="📥 Download Processed Excel",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_level6_summary",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"Error: {str(e)}")
            else:
                st.info("Please upload an Excel file to generate the summary table.")
        elif bucket_option == "SBF SALAD NEGATIVE ACCOUNT":
            st.subheader("SBF SALAD NEGATIVE ACCOUNT Email Blast File Uploader")
            uploaded_file = st.file_uploader(
                "📤 Choose an Excel file",
                type=["xlsx"],
                key="sbf_salad_uploader",
                help="Upload an Excel (.xlsx) with columns: Email, Name, Product Type, Client Name, Account No., Financing/Card No."
            )
            if uploaded_file is not None:
                st.session_state.uploaded_file = uploaded_file
                st.success("File uploaded successfully!")
            if st.session_state.uploaded_file is not None:
                if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                    st.session_state.uploaded_file = None
                    st.rerun()
            if st.session_state.uploaded_file is not None:
                try:
                    df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl')
                    # Validate Email column for '@' symbol
                    initial_row_count_email = len(df)
                    df['Email'] = df['Email'].astype(str)
                    df = df[df['Email'].str.contains('@', na=False)]
                    if len(df) < initial_row_count_email:
                        st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
                    required_columns = ['Email', 'Name', 'Product Type', 'Client Name', 'Account No.', 'Financing/Card No.']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        st.error(f"Missing required columns in the uploaded file: {', '.join(missing_columns)}")
                    else:
                        df = df[required_columns]
                        summary_df = pd.DataFrame({
                            'Email': df['Email'],
                            '{{chname}}': df['Name'],
                            '{{ID}}': 'SDCO',
                            'Client Name': df['Client Name'],
                            'Account No.': df['Account No.'],
                            'Financing/Card No.': df['Financing/Card No.']
                        })
                        text_columns = ['Email', '{{chname}}', '{{ID}}', 'Client Name', 'Account No.', 'Financing/Card No.']
                        for col in text_columns:
                            summary_df[col] = summary_df[col].astype(str)
                        st.write("### Processed Data")
                        st.dataframe(summary_df, use_container_width=True)
                        st.write("### Summary")
                        st.write("Note: Values in Account No. and Financing/Card No. are preserved exactly as uploaded.")
                        total_records = len(summary_df)
                        unique_emails = summary_df['Email'].nunique()
                        unique_names = summary_df['{{chname}}'].nunique()
                        unique_ids = summary_df['{{ID}}'].nunique()
                        unique_clients = summary_df['Client Name'].nunique()
                        unique_accounts = summary_df['Account No.'].nunique()
                        unique_financing = summary_df['Financing/Card No.'].nunique()
                        st.write(f"- **Total Records**: {total_records}")
                        st.write(f"- **Unique Emails**: {unique_emails}")
                        st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                        st.write(f"- **Unique IDs ({{ID}})**: {unique_ids}")
                        st.write(f"- **Unique Client Names**: {unique_clients}")
                        st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                        st.write(f"- **Unique Financing/Card Numbers**: {unique_financing}")
                        output = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Summary"
                        headers = list(summary_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(summary_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                        wb.save(output)
                        output.seek(0)
                        today = datetime.now().strftime("%B %d %Y")
                        file_name = f"SBF Salad Negative Account Email blasting {today}.xlsx"
                        st.download_button(
                            label="📥 Download Processed Excel",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_sbf_salad_summary",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"An error occurred while processing the file: {str(e)}")
            else:
                st.info("Please upload an Excel file to generate the summary table.")
        elif bucket_option == "SBF PL NEGATIVE ACCOUNTS":
            st.subheader("SBF PL NEGATIVE ACCOUNTS Email Blast File Uploader")
            uploaded_file = st.file_uploader(
                "📤 Choose an Excel file",
                type=["xlsx"],
                key="sbf_pl_uploader",
                help="Upload an Excel (.xlsx) with columns: Email, Name, Product Type, Client Name, Account No., Financing/Card No."
            )
            if uploaded_file is not None:
                st.session_state.uploaded_file = uploaded_file
                st.success("File uploaded successfully!")
            if st.session_state.uploaded_file is not None:
                if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                    st.session_state.uploaded_file = None
                    st.rerun()
            if st.session_state.uploaded_file is not None:
                try:
                    df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl')
                    # Validate Email column for '@' symbol
                    initial_row_count_email = len(df)
                    df['Email'] = df['Email'].astype(str)
                    df = df[df['Email'].str.contains('@', na=False)]
                    if len(df) < initial_row_count_email:
                        st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
                    required_columns = ['Email', 'Name', 'Product Type', 'Client Name', 'Account No.', 'Financing/Card No.']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        st.error(f"Missing required columns in the uploaded file: {', '.join(missing_columns)}")
                    else:
                        df = df[required_columns]
                        summary_df = pd.DataFrame({
                            'Email': df['Email'],
                            '{{chname}}': df['Name'],
                            '{{ID}}': 'PDCO',
                            'Client Name': df['Client Name'],
                            'Account No.': df['Account No.'],
                            'Financing/Card No.': df['Financing/Card No.']
                        })
                        text_columns = ['Email', '{{chname}}', '{{ID}}', 'Client Name', 'Account No.', 'Financing/Card No.']
                        for col in text_columns:
                            summary_df[col] = summary_df[col].astype(str)
                        st.write("### Processed Data")
                        st.dataframe(summary_df, use_container_width=True)
                        st.write("### Summary")
                        st.write("Note: Values in Account No. and Financing/Card No. are preserved exactly as uploaded.")
                        total_records = len(summary_df)
                        unique_emails = summary_df['Email'].nunique()
                        unique_names = summary_df['{{chname}}'].nunique()
                        unique_ids = summary_df['{{ID}}'].nunique()
                        unique_clients = summary_df['Client Name'].nunique()
                        unique_accounts = summary_df['Account No.'].nunique()
                        unique_financing = summary_df['Financing/Card No.'].nunique()
                        st.write(f"- **Total Records**: {total_records}")
                        st.write(f"- **Unique Emails**: {unique_emails}")
                        st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                        st.write(f"- **Unique IDs ({{ID}})**: {unique_ids}")
                        st.write(f"- **Unique Client Names**: {unique_clients}")
                        st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                        st.write(f"- **Unique Financing/Card Numbers**: {unique_financing}")
                        output = BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Summary"
                        headers = list(summary_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(summary_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                        wb.save(output)
                        output.seek(0)
                        today = datetime.now().strftime("%B %d %Y")
                        file_name = f"SBF PL Negative Account Email blasting {today}.xlsx"
                        st.download_button(
                            label="📥 Download Processed Excel",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_sbf_pl_summary",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"An error occurred while processing the file: {str(e)}")
            else:
                st.info("Please upload an Excel file to generate the summary table.")
    elif st.session_state.button3_clicked:
        st.subheader("LIVE INBOUND MONITORING")
        st.markdown(
            """
            <iframe src="https://spmadridlaw.sg.larksuite.com/share/base/dashboard/shrlgmGDFf4zcgqMR1vVl9044Nh" 
            class="dashboard-iframe" 
            width="100%" 
            height="600px" 
            frameborder="0" 
            allow="fullscreen">
            </iframe>
            """,
            unsafe_allow_html=True
        )

    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="footer">Viber Blast Uploader v1.3 | Jul 28, 2025 10:43 AM PST</div>', unsafe_allow_html=True)